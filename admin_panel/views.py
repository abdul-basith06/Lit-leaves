from django.http import Http404, HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from django.shortcuts import render,redirect
import openpyxl
import openpyxl.styles
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from .helpers import render_to_pdf
from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models import Sum, Count, Subquery, OuterRef
from django.db.models.functions import TruncWeek, TruncMonth, TruncDay, ExtractWeek
from .forms import CouponForm
from .models import *
from django.db.models import Sum, F, DecimalField, Value, Case, When
from userauths.models import User
from collections import defaultdict
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.views.decorators.cache import never_cache
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.shortcuts import get_object_or_404
from django.core.exceptions import ValidationError
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from shop.models import *


from django.contrib.auth.decorators import user_passes_test

def superuser_required(view_func):
    """
    Decorator for views that checks that the user is a superuser.
    """
    actual_decorator = user_passes_test(
        lambda u: u.is_active and u.is_superuser,
        login_url='admin_panel:admin_login',
    )
    return actual_decorator(view_func)

def admin_login(request):
    if request.user.is_authenticated and request.user.username == 'basi':
        return redirect('admin_panel:admin_dash')
    if request.method == 'POST':
        admin_name = request.POST['username']
        password = request.POST['password']
        if admin_name == 'basi' and password == '12369':
            user = authenticate(request, username=admin_name, password=password)

            if user is not None:
                login(request, user)
                return redirect('admin_panel:admin_dash')
        else:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'admin_panel/admin_login.html')
    
@never_cache
@superuser_required
def admin_logout(request):
    if request.user.is_authenticated:
        request.session.flush()
    return redirect('admin_panel:admin_login')
       
@superuser_required
def unlist_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    product.is_active = False 
    product.save()
    return redirect('admin_panel:products')

@superuser_required
def list_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    product.is_active = True 
    product.save()
    return redirect('admin_panel:products')


@superuser_required
def update_products(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        price = request.POST.get('price')
        author = request.POST.get('author')
        category_id = request.POST.get('category')
        
        product.name = name
        product.description = description
        product.price = price
        product.author = author
        product.category_id = category_id
        
        new_images = request.FILES.getlist('new_images')
        delete_images = request.POST.getlist('delete_images')

        for new_image in new_images:
            product_image = ProductImage(image=new_image, product=product)
            product_image.save()

        for image_id in delete_images:
            product_image = ProductImage.objects.get(id=image_id)
            product_image.delete()
            
        product.save()

        return redirect('admin_panel:products')

    context = {
        'product': product,
    }

    return render(request, 'admin_panel/edit_product.html', context)

@superuser_required
def edit_products(request, product_id):
    product = Product.objects.get(pk=product_id)
    cat2 = Categories.objects.all()
    context = {
        'product': product,
        'cat2': cat2,
    }
    return render(request, 'admin_panel/edit_products.html',context)

@superuser_required
def addd_products(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        author = request.POST.get('author')
        price = request.POST.get('price')
        category_id = request.POST.get('category')
        
        if int(price) <= 0:
            messages.error(request, "Enter a valid price !")
            return redirect('admin_panel:add_products')
      
       
        category = Categories.objects.get(pk=category_id)

        product = Product.objects.create(
            name=name,
            description=description,
            author=author,
            price=price,
            category=category,
        )

        images = request.FILES.getlist('image')  
        for image in images:
            ProductImage.objects.create(product=product, image=image)

        return redirect('admin_panel:products')
    else:
        pass

    
@superuser_required    
def add_products(request):
    all_cat = Categories.objects.filter(is_active=True)
    context = {
        'all_cat' : all_cat
    } 
    return render(request, 'admin_panel/add_products.html',context)

@superuser_required
def product_variation(request, product_id):
    product = Product.objects.get(pk=product_id)
    languages = Language.objects.all() 
    context = {
        'product': product,
        'languages' : languages,
    }
    
    return render(request, 'admin_panel/product_variation.html', context)

@superuser_required
def edit_stock_variation(request, variation_id):
    try:
        product_variation = ProductLanguageVariation.objects.get(pk=variation_id)
    except ProductLanguageVariation.DoesNotExist:
        raise Http404("Product Variation does not exist")
    
    if request.method == 'POST':
        stock = request.POST.get('stock', 0)
        language_id = request.POST.get('language')
        try:
            stock = int(stock)
            if stock < 0:
                messages.error(request, 'Stock cannot be negative.')
                return redirect('admin_panel:product_variation', product_id=product_variation.product.id)
        except ValueError:
            messages.error(request, 'Stock must be a valid integer.')
            return redirect('admin_panel:product_variation', product_id=product_variation.product.id)
        
        product_variation.stock = stock
        product_variation.language_id = language_id
        product_variation.save()
        return redirect('admin_panel:product_variation', product_id=product_variation.product.id) 
    
    context = {
        'product_variation': product_variation,
    }
    
    return render(request, 'admin_panel/product_variation.html', context)


@superuser_required
def add_variant(request, product_id):
    if request.method == 'POST':
        product = Product.objects.get(pk=product_id)
        language_id = request.POST['language']
        stock = request.POST['stock']
        try:
            stock = int(stock)
            if stock < 0:
                messages.error(request, 'Stock cannot be negative.')
                return redirect('admin_panel:product_variation', product_id=product_id)
            
            existing_variant = ProductLanguageVariation.objects.filter(product=product, language_id=language_id).first()
            if existing_variant:
                messages.warning(request, f'A variant with {existing_variant.language.name} already exists for this product.')
                return redirect('admin_panel:product_variation', product_id=product_id)
            else:
                ProductLanguageVariation.objects.create(product=product, language_id=language_id, stock=stock)
                messages.success(request, 'Variant added successfully.')
                return redirect('admin_panel:product_variation', product_id=product_id)
        except ValueError:
            messages.error(request, 'Stock must be a valid integer.')

    return render(request, 'admin_panel/product_variation.html')  



@superuser_required
def products(request):
    search_query = request.GET.get('key')  
    pro1 = Product.objects.all()

    if search_query:
        pro1 = pro1.filter(Q(name__icontains=search_query))
        

    context = {
        'pro1': pro1,
       'search_query': search_query,
        
    }
    return render(request, 'admin_panel/products.html', context)



@superuser_required
def category(request):
    search_query = request.GET.get('key')  
    cat1 = Categories.objects.all()

    if search_query:
        cat1 = cat1.filter(name__icontains=search_query)
     
    context = {
        'cat1': cat1,
        'search_query': search_query, 
    }
    return render(request, 'admin_panel/category.html', context)



@superuser_required
def add_category(request):
    if request.method == 'POST':
        category_name = request.POST.get('category_name')
        icon_file = request.FILES.get('category_icon')

        if category_name and icon_file:
            category = Categories(name=category_name)

            icon_path = default_storage.save('images/category_icons/' + icon_file.name, ContentFile(icon_file.read()))
            category.icon = icon_path
            category.save()

            return redirect('admin_panel:category')

    return render(request, 'add_category_modal.html')


@superuser_required
def add_language(request):
    if request.method == 'POST':
        language_name = request.POST.get('language_name')

        if language_name:
            language = Language(name=language_name)

            language.save()
            return redirect('admin_panel:category')


@superuser_required
def block_category(request, category_id):
    category = Categories.objects.get(pk=category_id)
    category.is_active = False
    category.save()
    Product.objects.filter(category=category).update(is_active=False)
    return redirect('admin_panel:category')

@superuser_required
def unblock_category(request, category_id):
    category = Categories.objects.get(pk=category_id)
    category.is_active = True
    category.save()
    Product.objects.filter(category=category).update(is_active=True)
    return redirect('admin_panel:category')

@superuser_required
def edit_category(request, category_id):
    category = Categories.objects.get(pk=category_id)
    
    context = {
        'category': category,
    }
    
    return render(request, 'edit_category_modal.html', context)

@superuser_required
def update_category(request, category_id):
    category = Categories.objects.get(pk=category_id)
    
    if request.method == 'POST':
        category.name = request.POST['category_name']
        if 'category_icon' in request.FILES:
            icon_file = request.FILES['category_icon']
            icon_path = default_storage.save('images/category_icons/' + icon_file.name, ContentFile(icon_file.read()))
            category.icon = icon_path
        category.save()
        return redirect('admin_panel:category') 

    context = {
        'category': category,
    }
    
    return render(request, 'edit_category_modal.html', context)



@superuser_required
def delete_category(request, category_id):
    if request.method == 'POST':
        category = get_object_or_404(Categories, pk=category_id)

        Product.objects.filter(category=category).update(is_active=False)

        category.delete()
        messages.success(request, 'Category and associated products deactivated successfully.')
    else:
        messages.error(request, 'Invalid request method')

    return redirect('admin_panel:category')
 
@superuser_required
def block_user(request, user_id):
    
    if request.user.id == int(user_id):
        logout(request)
        
    user = User.objects.get(pk=user_id)
    user.is_active = False
    user.save()
    return redirect('admin_panel:user_management')

@superuser_required
def unblock_user(request, user_id):
    user = User.objects.get(pk=user_id)
    user.is_active = True
    user.save()
    return redirect('admin_panel:user_management')



@superuser_required
def user_management(request):
    search_query = request.GET.get('key')
    cus = User.objects.all()
    
    if search_query:
        cus = cus.filter(Q(username__icontains=search_query) | Q(email__icontains=search_query))

    context = {
        'cus': cus,
        'search_query': search_query,
    }
    return render(request, 'admin_panel/user_manage.html', context)


@superuser_required
def admin_dash(request):
    all_orders = Order.objects.filter(orderitem__delivery_status='D', complete=True).distinct()
    all_variations = ProductLanguageVariation.objects.all()
    all_order_items = OrderItem.objects.filter(delivery_status='D')
    delivered_order_items = OrderItem.objects.filter(delivery_status='D')

    
    filter_type = request.GET.get('filter_type', 'all')  

 
    if filter_type == 'day':
        start_date = datetime.now() - timedelta(days=1)
    elif filter_type == 'week':
        start_date = datetime.now() - timedelta(weeks=1)
    elif filter_type == 'month':
        start_date = datetime.now() - timedelta(weeks=4)  
    elif filter_type == 'year':
        start_date = datetime.now() - timedelta(weeks=52)  
    else:
        start_date = None  
    
    if start_date:
         all_orders = all_orders.filter(
            date_ordered__gte=start_date,
            orderitem__delivery_status='D'
        ).distinct()
         
         delivered_order_items = OrderItem.objects.filter(
            order__in=all_orders,
            delivery_status='D'
        )

   
    total_revenue = sum(order_item.get_total for order_item in delivered_order_items)
    
    total_sales = sum(
    sum(item.quantity for item in order.orderitem_set.filter(delivery_status='D'))
    for order in all_orders
)

    total_stock = sum(variation.stock for variation in all_variations)
    
    
    cod_orders = all_orders.filter(payment_method='COD')
    online_orders = all_orders.filter(payment_method='RAZ') 
    wallet_orders = all_orders.filter(payment_method='WAL')

    cod_count = cod_orders.count()
    online_count = online_orders.count()
    wallet_count = wallet_orders.count()

    cod_total = sum(order.get_cart_total for order in cod_orders)
    online_total = sum(order.get_cart_total for order in online_orders)
    wallet_total = sum(order.get_cart_total for order in wallet_orders)

    
    context = {
        'total_revenue': total_revenue,
        'total_sales' : total_sales,
        'total_stock':total_stock,
        'all_orders':all_orders,
        'all_order_items':all_order_items,
        'cod_count': cod_count,
        'online_count': online_count,
        'wallet_count': wallet_count,
        'cod_total': cod_total,
        'online_total': online_total,
        'wallet_total': wallet_total,
        'filter_type': filter_type,  
       

    }
    return render(request, 'admin_panel/admin_dash.html', context)

@superuser_required
def order_management(request):
    order = Order.objects.filter(complete=True)
    context = {
        'order' : order,
    }
    return render(request, 'admin_panel/orders.html', context)

@superuser_required
def manage_order(request, order_id, orderitem_id):
    order = get_object_or_404(Order, id=order_id)
    order_item = get_object_or_404(OrderItem, id=orderitem_id)


    context = {
        'order': order,
        'order_item': order_item,
    }

    return render(request, 'admin_panel/manage_order.html', context)

@superuser_required
def cancel_order(request, order_item_id):
    order_item = get_object_or_404(OrderItem, id=order_item_id)
    
    order_item.variation.stock += order_item.quantity
    order_item.variation.save() 
    
    order_item.delivery_status = 'CN'
    order_item.save()
    
    if order_item.order.payment_method in ['WAL', 'RAZ']:  
        refund_amount = order_item.get_total() if callable(order_item.get_total) else order_item.get_total

        user = order_item.order.customer
        user_wallet = Wallet.objects.get(user=user)
        order = order_item.order

        if order.payment_method == 'WAL' or order.payment_method == 'RAZ':
            user_wallet.balance += refund_amount
            user_wallet.save()

    messages.success(request, 'Order canceled successfully.')

    return redirect('admin_panel:order_management')

@superuser_required
def update_order_status(request, order_item_id):
    if request.method == 'POST':
        delivery_status = request.POST.get('delivery_status')
        order_item = OrderItem.objects.get(id=order_item_id)
        
        if delivery_status == 'CN'and order_item.variation:
            order_item.variation.stock += order_item.quantity
            order_item.variation.save() 
            
            if order_item.order.payment_method in ['WAL', 'RAZ']:  
                refund_amount = order_item.get_total() if callable(order_item.get_total) else order_item.get_total
                
                user = order_item.order.customer
                user_wallet = Wallet.objects.get(user=user)
                order = order_item.order
                
                if order.payment_method == 'WAL' or order.payment_method == 'RAZ':
                    user_wallet.balance += refund_amount
                    user_wallet.save()

        order_item.delivery_status = delivery_status
        order_item.save()

        return redirect('admin_panel:order_management')  

    return render(request, 'admin_panel/order_management.html')

@superuser_required
def coupons(request):
    coupon = Coupon.objects.all()
    context = {
        'coupon':coupon,
    }
    return render(request, 'admin_panel/coupons.html',context)

@superuser_required
def add_coupons(request):
    form = CouponForm()

    if request.method == 'POST':
        form = CouponForm(request.POST)
        if form.is_valid():
            coupon = form.save()
            messages.success(request, 'Coupon added successfully.')
            return redirect('admin_panel:coupons')
        else:
            messages.error(request, 'Error adding the coupon. Please check the form data.')
    context = {
        'form':form,
    }        
    return render(request, 'admin_panel/add_coupons.html',context)

@superuser_required
def edit_coupon(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    form = CouponForm(instance=coupon)
    context = {
        'form':form,
        'coupon_id':coupon_id,
    }
    return render(request, 'admin_panel/edit_coupon.html',context)

@superuser_required
def update_coupon(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    if request.method == 'POST':
        form = CouponForm(request.POST, instance=coupon)
        if form.is_valid():
            form.save()
            messages.success(request, 'Coupon edited successfully.')
            return redirect('admin_panel:coupons')
    else:
        form = CouponForm(instance=coupon)
    context = {
        'form':form,
        'coupon_id':coupon_id,
    }
    return render(request, 'admin_panel/edit_coupon.html',context)

@superuser_required
def delete_coupon(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    if request.method == 'POST':
        coupon.delete()
        messages.success(request, 'Coupon deleted successfully.')
        return redirect('admin_panel:coupons')
    
    return render(request, 'admin_panel/coupons.html')
   

@superuser_required    
def sales_report(request):
    all_orders = Order.objects.all()
   
    start_date_str = '1900-01-01'
    end_date_str = '9999-12-31'
    if request.method == 'POST':
        start_date_str = request.POST.get('start', '1900-01-01')
        end_date_str = request.POST.get('end', '9999-12-31')
   

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

    if start_date is not None and end_date is not None:
        delivered_order_items = OrderItem.objects.filter(
            delivery_status='D',
            order__complete=True,
            order__date_ordered__gte=start_date,
            order__date_ordered__lte=end_date
        )
    else:
        delivered_order_items = OrderItem.objects.filter(
            delivery_status='D',
            order__complete=True
        )


    sold_items = delivered_order_items.values('product__name').annotate(total_sold=Sum('quantity'))

    total_sales = sum(order.get_cart_total for order in all_orders)
    total_profit = int(sum(order_item.product.price * 0.3 for order_item in delivered_order_items))

    most_sold_products = (
        delivered_order_items.values('product__name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:6]
    )

    product_profit_data = []

    for sold_stock_item in delivered_order_items:
        product_price = sold_stock_item.product.price
        total_revenue = sold_stock_item.get_total 
        profit = total_revenue - (total_revenue * 0.7)
        profit = int(profit)
        product_profit_data.append({
            'product_name': sold_stock_item.product.name,
            'total_sold': sold_stock_item.quantity,
            'profit': profit,
        })

    context = {
        'total_sales': total_sales,
        'most_sold_products': most_sold_products,
        'total_profit': total_profit,
        'sold_items': sold_items,
        'stock_items': product_profit_data,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else None,
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else None,
    }

    return render(request, 'admin_panel/sales_report.html', context)
   
def sales_report_pdf(request):
    
    all_orders = Order.objects.all()
    start_date_str = request.GET.get('start', '1900-01-01')
    end_date_str = request.GET.get('end', '9999-12-31')

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
      
    delivered_order_items = OrderItem.objects.filter(
            delivery_status='D',
            order__complete=True,
            order__date_ordered__gte=start_date,
            order__date_ordered__lte=end_date
        )
    total_revenue = sum(order_item.get_total for order_item in delivered_order_items)
    total_sales = sum(order_item.quantity for order_item in delivered_order_items)
    total_profit = sum(order_item.product.price * 0.3 for order_item in delivered_order_items)
    
    all_orders = Order.objects.filter(orderitem__in=delivered_order_items).distinct()
    
    sold_items = [{
        'product_name': item.product.name,
        'total_sold': item.quantity,
        'profit': int(item.get_total - (item.get_total * 0.7)),
    } for item in delivered_order_items]
    sold_items = sorted(sold_items, key=lambda x: x['total_sold'], reverse=True)

    context = {
        'start_date':start_date,
        'end_date':end_date,
        'total_sales': total_sales,
        'total_revenue':total_revenue,
        'total_profit':total_profit,
        'sold_items':sold_items,
        'all_orders':all_orders,
        'delivered_order_items':delivered_order_items,
        
    }

    pdf = render_to_pdf('admin_panel/sales_report_pdf.html', context)

    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = "Sales Report.pdf"
        content = f'attachment; filename="{filename}"'
        response['Content-Disposition'] = content
        return response

    return HttpResponse('Failed to generate PDF')

def sales_report_excel(request):
    try:
        start_date_str = request.GET.get('start', '1900-01-01')
        end_date_str = request.GET.get('end', '9999-12-31')
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
        delivered_order_items = OrderItem.objects.filter(
            delivery_status='D',
            order__complete=True,
            order__date_ordered__gte=start_date,
            order__date_ordered__lte=end_date
        )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Total Revenue'
        worksheet['B1'] = sum(order_item.get_total for order_item in delivered_order_items)
        worksheet['A2'] = 'Total Sales'
        worksheet['B2'] = sum(order_item.quantity for order_item in delivered_order_items)
        worksheet['A3'] = 'Total Profit'
        worksheet['B3'] = sum(order_item.product.price * 0.3 for order_item in delivered_order_items)

        row_num = 5
        headers = ['Order ID', 'Date Ordered', 'Total Amount', 'Products']
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=row_num, column=col_num, value=header)

        row_num += 1
        date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')
        for order_item in delivered_order_items:
            worksheet.cell(row=row_num, column=1, value=order_item.order.id)

            date_ordered_naive = order_item.order.date_ordered.astimezone(timezone.utc).replace(tzinfo=None)
            worksheet.cell(row=row_num, column=2, value=date_ordered_naive)
            worksheet.cell(row=row_num, column=2).style = date_style

            worksheet.cell(row=row_num, column=3, value=order_item.order.get_cart_total)

            products_list = [f"{item.product.name} ({item.quantity} units) - ${item.get_total}" for item in order_item.order.orderitem_set.all()]
            products_str = '\n'.join(products_list)
            worksheet.cell(row=row_num, column=4, value=products_str)

            row_num += 1

        for i in range(1, 4):
            cell = worksheet.cell(row=i, column=2)
            cell.font = openpyxl.styles.Font(bold=True)

        header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        header_alignment = openpyxl.styles.Alignment(horizontal='center')
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num - len(delivered_order_items) - 1, column=col_num)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = openpyxl.styles.PatternFill(start_color='007bff', end_color='007bff', fill_type='solid')

        data_alignment = openpyxl.styles.Alignment(wrap_text=True)
        for row_num in range(row_num - len(delivered_order_items), row_num):
            for col_num in range(1, 5):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = data_alignment
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
                
        for col_num, header in enumerate(headers, 1):
            max_length = len(header)
            for row_num in range(2, row_num):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                try:
                    if len(str(cell_value)) > max_length:
                         max_length = len(cell_value)
                except:
                        pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Sales_Report.xlsx'
        workbook.save(response)

        return response

    except Exception as e:
        print(f"Error generating Excel file: {e}")
        return HttpResponse("Failed to generate Excel file", status=500)
   

@superuser_required
@require_GET
def get_sales_data(request, period):
    if period == 'week':
        
        start_date = timezone.now().date() - timezone.timedelta(days=6)
        order_items = OrderItem.objects.filter(order__date_ordered__gte=start_date)
        data = (
            order_items.annotate(day=TruncDay('order__date_ordered'))
            .values('day')
            .annotate(total=Sum(F('quantity') * F('product__price')))
            .order_by('day')
        )
        labels = [item['day'].strftime('%A') for item in data]
    elif period == 'month':
        start_date = timezone.now().date() - timezone.timedelta(days=30)
        order_items = OrderItem.objects.filter(order__date_ordered__gte=start_date)
        data = (
        order_items.annotate(day=TruncDay('order__date_ordered'))
        .values('day')
        .annotate(total=Sum(F('quantity') * F('product__price')))
        .order_by('day')
    )
        labels = [item['day'].strftime('%Y-%m-%d') for item in data]
    elif period == 'year':
        start_date = timezone.now().date() - timezone.timedelta(days=365)
        order_items = OrderItem.objects.filter(order__date_ordered__gte=start_date)
        data = (
            order_items.annotate(month=TruncMonth('order__date_ordered'))
            .values('month')
            .annotate(total=Sum(F('quantity') * F('product__price')))
            .order_by('month')
        )
        labels = [f"{item['month'].strftime('%B')}" for item in data]
    else:
        return JsonResponse({'error': 'Invalid period'})

    
    
    sales_data = [item['total'] for item in data]

    return JsonResponse({'labels': labels, 'data': sales_data})






















# def sales_portfolio(request):
#     print("coming into this view>>>>>>>>>>>>>>>>>>>>>")
#     time_period = request.POST.get('date', 'week') 

#     end_date = timezone.now()
#     if time_period == 'week':
#         start_date = end_date - timezone.timedelta(days=7)
#     elif time_period == 'month':
#         start_date = end_date - timezone.timedelta(days=30)
#     elif time_period == 'year':
#         start_date = end_date - timezone.timedelta(days=365)
#     else:
#         # Default to 'week' if an invalid time period is provided
#         start_date = end_date - timezone.timedelta(days=7)

#     # Fetch orders within the selected time period
#     orders_within_period = Order.objects.filter(date_ordered__range=[start_date, end_date])

#     # Group orders by date
#     sales_data = orders_within_period.values('date_ordered')

#     # Calculate total sales for each date
#     sales_data = sales_data.annotate(
#         total_sales=Sum(F('orderitem__product__price') * F('orderitem__quantity'), output_field=DecimalField())
#     )

#     # Extract labels (dates) and data (total sales)
#     labels = [order['date_ordered'].strftime('%Y-%m-%d') for order in sales_data]
#     data = [order['total_sales'] if order['total_sales'] is not None else 0 for order in sales_data]

#     # Calculate overall total sales
#     overall_total_sales = sum(data)
    
#     print('Labels:', labels)
#     print('Data:', data)
#     print('Overall Total Sales:', overall_total_sales)

#     return JsonResponse({'labels': labels, 'data': data, 'overall_total_sales': overall_total_sales})



    